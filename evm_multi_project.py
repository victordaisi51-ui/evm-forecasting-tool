"""
EVM Multi-Project Tool
-----------------------
Reads ONE CSV file that contains data for MANY projects (identified by
'project_id'), and calculates CPI/SPI/EAC for EACH project separately.

NEW CONCEPT: groupby
---------------------
Right now, your CSV has 38 rows total, but they represent 10 SEPARATE
projects mixed together in one file. `groupby("project_id")` tells pandas:
"split this one big table into 10 smaller tables, one per project ID,
and let me process each one on its own."
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from evm_basic import calculate_evm


# -----------------------------------------------------------------------
# STEP 1: Load the file, letting the user specify which file to load
# -----------------------------------------------------------------------
def load_data(csv_path):
    try:
        df = pd.read_csv(csv_path, parse_dates=["period"])
    except FileNotFoundError:
        print(f"Could not find a file at: {csv_path}")
        raise

    # Validate the columns we absolutely need are present.
    required_columns = {
        "project_id", "project_name", "period",
        "budgeted_cost", "percent_planned",
        "percent_complete", "actual_cost",
    }
    missing = required_columns - set(df.columns)
    if missing:
        raise ValueError(f"CSV is missing required columns: {missing}")

    return df


# -----------------------------------------------------------------------
# STEP 2: Process EVERY project and collect a one-line summary for each
# -----------------------------------------------------------------------
def summarize_all_projects(df):
    summary_rows = []

    # df.groupby("project_id") hands us (project_id, sub_dataframe) pairs,
    # one at a time, in a for loop -- same "for" concept you already know,
    # just looping over GROUPS instead of individual numbers.
    for project_id, project_df in df.groupby("project_id"):
        bac = project_df["budgeted_cost"].sum()

        result = calculate_evm(project_df, bac)

        # .iloc[-1] means "the LAST row" -- i.e. the most recent period
        latest = result.iloc[-1]

        summary_rows.append({
            "project_id": project_id,
            "project_name": latest["project_name"] if "project_name" in latest else project_df["project_name"].iloc[0],
            "CPI": round(latest["CPI"], 2),
            "SPI": round(latest["SPI"], 2),
            "EAC": round(latest["EAC"], 2),
            "status": "AT RISK" if (latest["CPI"] < 1 or latest["SPI"] < 1) else "HEALTHY",
        })

    summary_df = pd.DataFrame(summary_rows)

    # Sort so the worst-performing projects show up first -- most useful
    # for a project manager scanning for problems.
    summary_df = summary_df.sort_values(by="CPI")

    return summary_df


# -----------------------------------------------------------------------
# STEP 3: Export the summary as a formatted Excel file
# -----------------------------------------------------------------------
def export_to_excel(summary_df, output_path):
    """
    Builds a proper .xlsx file (not just a CSV) with:
    - Bold headers
    - Red fill for AT RISK projects, green fill for HEALTHY ones
    - Sensible column widths so nothing looks cramped
    """
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Portfolio Summary"

    headers = list(summary_df.columns)
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    red_fill = PatternFill("solid", start_color="FFC7CE")
    green_fill = PatternFill("solid", start_color="C6EFCE")

    for _, row in summary_df.iterrows():
        sheet.append(list(row))
        excel_row = sheet.max_row
        fill = red_fill if row["status"] == "AT RISK" else green_fill
        for cell in sheet[excel_row]:
            cell.fill = fill

    widths = [12, 26, 8, 8, 14, 12]
    for i, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + i)].width = width

    wb.save(output_path)


# -----------------------------------------------------------------------
# STEP 4: Tie it together
# -----------------------------------------------------------------------
if __name__ == "__main__":
    csv_path = input("Enter the path to your CSV file (or press Enter for the default sample): ").strip()

    if csv_path == "":
        csv_path = "data/multi_project.csv"

    try:
        df = load_data(csv_path)
    except (FileNotFoundError, ValueError) as e:
        print(f"\nCould not process the file: {e}")
    else:
        summary = summarize_all_projects(df)

        print("\n--- Portfolio Summary (worst performers first) ---\n")
        print(summary.to_string(index=False))

        summary.to_csv("portfolio_summary.csv", index=False)
        export_to_excel(summary, "portfolio_summary.xlsx")
        print("\nFull summary saved to: portfolio_summary.csv")
        print("Formatted Excel report saved to: portfolio_summary.xlsx")
